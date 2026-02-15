import json
import os
import time
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Any, Optional
import pandas as pd

from t_tech.invest import Client
from t_tech.invest.schemas import (
    GetBondEventsRequest,
    EventType,
    OperationType,
    MoneyValue,
    Quotation,
)
from t_tech.invest.exceptions import RequestError
from google.protobuf.timestamp_pb2 import Timestamp

# from shared.config import TOKEN

# Импорты для API Т-Банка
from t_tech.invest import Client
from t_tech.invest.schemas import GetAssetFundamentalsRequest
from t_tech.invest.schemas import InstrumentExchangeType, InstrumentStatus
from t_tech.invest.schemas import GetBondEventsRequest, EventType, MoneyValue, Quotation
from google.protobuf.timestamp_pb2 import Timestamp

TOKEN = "t.V4QVXUA5khrTJcMQNsCCDC3IfD94uJA5Yj_FpR8UfaMs3KxSY0tlIlSDe3ix6G7CcKYMbfQTNlLSWR2l1aHQjQ"
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(parent_dir, "data", "fundamentals_shares.xlsx")
bonds_path = os.path.join(parent_dir, "data", "bonds_data.xlsx")
os.makedirs(os.path.dirname(file_path), exist_ok=True)

file_path_json = os.path.join(parent_dir, "data", "shares.json")
bonds_json_path = os.path.join(parent_dir, "data", "bonds_data.json")

# Константы для пакетной обработки
CHUNK_SIZE = 30
DELAY_BETWEEN_CHUNKS = 3
DELAY_BETWEEN_REQUESTS = 0.1


def quotation_to_float(q: Quotation) -> float:
    """Преобразует Quotation в float"""
    if not q:
        return 0.0
    return q.units + q.nano / 1e9


def parse_nominal(nominal_str: str) -> float:
    """Парсит номинал из строки в число"""
    try:
        content = nominal_str.replace("MoneyValue(", "").rstrip(")")
        parts = content.split(", ")
        units = 0
        nano = 0
        for part in parts:
            if part.startswith("units"):
                units = int(part.split("=")[1])
            elif part.startswith("nano"):
                nano = int(part.split("=")[1])
        return units + nano / 1e9
    except:
        return 0.0


def parse_date(date_str: str) -> Optional[str]:
    """Парсит дату в формат YYYY-MM-DD"""
    if not date_str or date_str == "1970-01-01 00:00:00+00:00":
        return None
    try:
        if "+" in date_str:
            date_str = date_str.split("+")[0]
        dt = datetime.fromisoformat(date_str)
        return (
            dt.replace(tzinfo=None).strftime("%Y-%m-%d")
            if dt.tzinfo
            else dt.strftime("%Y-%m-%d")
        )
    except:
        return None


def get_coupon_rate(client, instrument_id: str) -> Optional[float]:
    """Получает ставку купона по инструменту"""
    try:
        request = GetBondEventsRequest(
            instrument_id=instrument_id,
        )
        response = client.instruments.get_bond_events(request=request)

        if response.events:
            for event in response.events:
                if event.coupon_interest_rate:
                    rate = quotation_to_float(event.coupon_interest_rate)
                    if rate > 0:
                        return round(rate, 2)
        return None

    except RequestError as e:
        print(f"      ⚠️ Лимит API, ждём 10 сек...")
        time.sleep(10)
        return None
    except:
        return None


def get_bonds_data():
    """Основная функция"""
    print("🚀 Загрузка данных по облигациям")
    print("=" * 60)

    # === 1. Загружаем JSON ===
    with open(bonds_json_path, "r", encoding="utf-8") as f:
        instruments = json.load(f)["instruments"]

    print(f"📊 Всего облигаций: {len(instruments)}")
    print(f"📦 Чанк: {CHUNK_SIZE} шт, интервал: {DELAY_BETWEEN_CHUNKS} сек")
    print("=" * 60)

    # === 2. Парсим базовые данные ===
    bonds = []
    instruments_to_query = []

    for bond in instruments:
        instrument_id = bond.get("figi") or bond.get("uid")

        bond_data = {
            "ticker": bond.get("ticker", ""),
            "name": bond.get("name", ""),
            "sector": bond.get("sector", ""),
            "currency": bond.get("currency", ""),
            "floating_coupon_flag": bond.get("floating_coupon_flag", None),
            "amortization_flag": bond.get("amortization_flag", None),
            "perpetual_flag": bond.get("perpetual_flag", None),
            "maturity_date": parse_date(bond.get("maturity_date", "")),
            "nominal": parse_nominal(bond.get("nominal", "")),
            "risk_level": bond.get("risk_level", ""),
            "coupon_rate": None,
        }
        bonds.append(bond_data)

        if instrument_id:
            instruments_to_query.append(
                {
                    "ticker": bond_data["ticker"],
                    "instrument_id": instrument_id,
                    "data": bond_data,
                }
            )

    print(f"🎯 Будем запрашивать: {len(instruments_to_query)} облигаций")

    # === 3. Обрабатываем чанками ===
    chunks = [
        instruments_to_query[i : i + CHUNK_SIZE]
        for i in range(0, len(instruments_to_query), CHUNK_SIZE)
    ]

    success = 0

    with Client(TOKEN) as client:
        for chunk_idx, chunk in enumerate(chunks, 1):
            print(f"\n🔄 Чанк {chunk_idx}/{len(chunks)} ({len(chunk)} шт)")

            for item_idx, item in enumerate(chunk, 1):
                ticker_short = (
                    item["ticker"][:12] + ".."
                    if len(item["ticker"]) > 12
                    else item["ticker"]
                )
                name_short = (
                    item["data"]["name"][:25] + ".."
                    if len(item["data"]["name"]) > 25
                    else item["data"]["name"]
                )

                print(
                    f"   [{item_idx:2d}/{len(chunk)}] {ticker_short:12} — {name_short:25}",
                    end=" ",
                )

                rate = get_coupon_rate(client, item["instrument_id"])

                if rate:
                    item["data"]["coupon_rate"] = rate
                    success += 1
                    print(f"✅ {rate}%")
                else:
                    item["data"]["coupon_rate"] = 0
                    print(f"❌")

                time.sleep(DELAY_BETWEEN_REQUESTS)

            if chunk_idx < len(chunks):
                print(f"   ⏳ Ожидание {DELAY_BETWEEN_CHUNKS} сек...")
                time.sleep(DELAY_BETWEEN_CHUNKS)

    # === 4. Сохраняем Excel ===
    df = pd.DataFrame(bonds)

    # Сортируем по дате погашения
    if "maturity_date" in df.columns:
        df = df.sort_values("maturity_date")

    df.to_excel(bonds_path, index=False, engine="openpyxl")

    print("\n" + "=" * 60)
    print(f"✅ Готово! Сохранено: {bonds_path}")
    print("=" * 60)

    # === 5. Простая статистика ===
    rates = df[df["coupon_rate"].notna()]

    print(f"\n📊 Статистика:")
    print(f"   Всего облигаций: {len(df)}")
    print(f"   Найдено ставок: {len(rates)} ({len(rates)/len(df)*100:.1f}%)")

    if not rates.empty:
        print(f"   Средняя ставка: {rates['coupon_rate'].mean():.2f}%")
        print(f"   Мин ставка: {rates['coupon_rate'].min():.2f}%")
        print(f"   Макс ставка: {rates['coupon_rate'].max():.2f}%")

        # Топ-5 секторов по доходности
        print(f"\n🏆 Топ-5 секторов по доходности:")
        sector_avg = (
            rates.groupby("sector")["coupon_rate"].agg(["mean", "count"]).round(2)
        )
        sector_avg = (
            sector_avg[sector_avg["count"] >= 5]
            .sort_values("mean", ascending=False)
            .head(5)
        )
        for sector, row in sector_avg.iterrows():
            print(f"   {sector:15}: {row['mean']}% ({row['count']} шт)")

    return df


def get_shares(token):
    """Получает список акций через API Т-Банка"""
    try:
        with Client(token) as client:
            print("[Т-Банк] Запрашиваю список акций...")
            instruments_response = client.instruments.shares()

            # Фильтрация (только рублевые)
            filtered_instruments = [
                instrument
                for instrument in instruments_response.instruments
                if instrument.currency.lower() == "rub"
            ]

            print(f"[Т-Банк] Получено рублевых акций: {len(filtered_instruments)}")

            # Подготавливаем данные для сохранения
            data_to_save = {
                "source": "T-Bank API",
                "generated_at": datetime.now().isoformat(),
                "instruments": [
                    {
                        "uid": instrument.uid,
                        "asset_uid": instrument.asset_uid,
                        "figi": instrument.figi,
                        "ticker": instrument.ticker,
                        "name": instrument.name,
                        "lot": instrument.lot,
                        "currency": instrument.currency,
                        "sector": getattr(instrument, "sector", "Не указан"),
                        "class_code": instrument.class_code,
                    }
                    for instrument in filtered_instruments
                ],
            }
            # Сохраняем в JSON
            with open(file_path_json, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, indent=4, ensure_ascii=False, default=str)

            print(f"✅ Данные сохранены в 'tbank_shares.json'")
            return True

    except Exception as e:
        print(f"❌ Ошибка при получении акций: {e}")
        return False


def get_fundamentals_assets_to_excel(token, excel_filename=file_path):
    """
    Получает фундаментальные показатели акций Т-Банк и сохраняет в Excel
    """
    try:
        # Загружаем список инструментов
        with open(file_path_json, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        instruments_list = json_data["instruments"]
        asset_uids = [
            inst["asset_uid"] for inst in instruments_list if inst.get("asset_uid")
        ]

        if not asset_uids:
            print("❌ Нет asset_uid для запроса")
            return

        print(f"[Т-Банк] Загружено {len(asset_uids)} активов для запроса")

        all_fundamentals = []
        chunk_size = 30  # Можно регулировать в зависимости от лимитов API

        # Обработка чанками
        for i in range(0, len(asset_uids), chunk_size):
            chunk = asset_uids[i : i + chunk_size]
            instruments_chunk = instruments_list[i : i + chunk_size]

            print(
                f"🔄 Пачка {i//chunk_size + 1}/{(len(asset_uids)-1)//chunk_size + 1} ({len(chunk)} активов)"
            )
            try:
                with Client(token) as client:
                    # КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: передаем assets как массив
                    request = GetAssetFundamentalsRequest(assets=chunk)
                    response = client.instruments.get_asset_fundamentals(
                        request=request
                    )
                    print(f"   Получено показателей: {len(response.fundamentals)}")

                    # Сопоставляем фундаментальные данные с инструментами
                    for fundamental in response.fundamentals:
                        # Ищем соответствующий инструмент по asset_uid
                        matching_instrument = next(
                            (
                                inst
                                for inst in instruments_chunk
                                if inst.get("asset_uid") == fundamental.asset_uid
                            ),
                            None,
                        )

                        if not matching_instrument:
                            continue

                        # Функция безопасного получения значений
                        def safe_get(field, default=None):
                            try:
                                val = getattr(fundamental, field, default)
                                if val is None:
                                    return default
                                # Обработка дат
                                if field.endswith("_date") and hasattr(
                                    val, "ToDatetime"
                                ):
                                    try:
                                        return val.ToDatetime().strftime("%Y-%m-%d")
                                    except:
                                        return default
                                return val
                            except AttributeError:
                                return default

                        # Форматирование чисел
                        def fmt_num(val):
                            if isinstance(val, (int, float)):
                                if val is None or val == 0:
                                    return ""
                                if abs(val) >= 1_000_000_000:
                                    return f"{val/1_000_000_000:.2f} млрд"
                                elif abs(val) >= 1_000_000:
                                    return f"{val/1_000_000:.2f} млн"
                                elif abs(val) >= 1_000:
                                    return f"{val/1_000:.1f} тыс"
                            return val

                        # Собираем строку данных
                        row = {
                            "Тикер": matching_instrument.get("ticker", ""),
                            "Название": matching_instrument.get("name", ""),
                            "Asset UID": matching_instrument.get("asset_uid", ""),
                            "FIGI": matching_instrument.get("figi", ""),
                            "Валюта": matching_instrument.get("currency", ""),
                            "Рыночная капитализация": fmt_num(
                                safe_get("market_capitalization")
                            ),
                            "EV": fmt_num(safe_get("total_enterprise_value_mrq")),
                            "Выручка": fmt_num(safe_get("revenue_ttm")),
                            "Чистая прибыль": fmt_num(safe_get("net_income_ttm")),
                            "EBITDA": fmt_num(safe_get("ebitda_ttm")),
                            "P/E": safe_get("pe_ratio_ttm"),
                            "P/B": safe_get("price_to_book_ttm"),
                            "P/S": safe_get("price_to_sales_ttm"),
                            "P/FCF": safe_get("price_to_free_cash_flow_ttm"),
                            "ROE": safe_get("roe"),
                            "ROA": safe_get("roa"),
                            "ROIC": safe_get("roic"),
                            "EV/EBITDA": safe_get("ev_to_ebitda_mrq"),
                            "EV/S": safe_get("ev_to_sales"),
                            "FCF": safe_get("free_cash_flow_ttm"),
                            "CAGR_Sales": safe_get(
                                "five_year_annual_revenue_growth_rate"
                            ),
                            "Averange_dividend_yield": safe_get(
                                "five_years_average_dividend_yield"
                            ),
                            "Averange_cagr_dividend_yield": safe_get(
                                "five_year_annual_dividend_growth_rate"
                            ),
                            "Current_ratio_mr": safe_get("current_ratio_mrq"),
                            "Payot Ratio": safe_get("dividend_payout_ratio_fy"),
                            "NPM": safe_get("net_margin_mrq"),
                            "Debt": safe_get("total_debt_mrq"),
                            "Debt/Capital": safe_get("total_debt_to_equity_mrq"),
                            "Net_Debt/EBITDA": safe_get("net_debt_to_ebitda"),
                            "Debt/EBITDA": safe_get("total_debt_to_ebitda_mrq"),
                            "EPS": safe_get("eps_ttm"),
                            "Дивидендная доходность": safe_get(
                                "dividend_yield_daily_ttm"
                            ),
                            "Бета": safe_get("beta"),
                            "Дивиденд на акцию": safe_get("dividends_per_share"),
                        }
                        all_fundamentals.append(row)

                    print(f"   ✓ Обработано записей: {len(response.fundamentals)}")

            except Exception as e:
                print(f"   ❌ Ошибка в пачке: {e}")
                continue

        # Сохраняем в Excel
        if all_fundamentals:
            print(all_fundamentals)
            save_to_excel(all_fundamentals, excel_filename)
        else:
            print("⚠️ Не получено данных для сохранения")

    except Exception as e:
        print(f"❌ Общая ошибка: {e}")


def save_to_excel(data, filename):
    """Сохраняет данные в Excel с форматированием"""
    try:
        df = pd.DataFrame(data)

        if df.empty:
            print("❌ Нет данных для сохранения")
            return

        print(f"\n💾 Сохраняю {len(df)} записей в '{filename}'...")

        # Сохраняем в Excel
        df.to_excel(filename, index=False, engine="openpyxl")

        # Применяем форматирование
        format_excel_file(filename, df)

        print(f"✅ Данные успешно сохранены!")
        print(f"📊 Статистика: {len(df)} строк, {len(df.columns)} столбцов")

    except Exception as e:
        print(f"❌ Ошибка сохранения: {e}")


def format_excel_file(filename, df):
    """Форматирует Excel файл"""
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Стиль заголовков
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Форматируем заголовки
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_len = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
                except:
                    pass
            adjusted_width = min(max_len + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Добавляем фильтры
        ws.auto_filter.ref = ws.dimensions

        wb.save(filename)
        print("   ✓ Форматирование применено")

    except Exception as e:
        print(f"   ⚠️ Ошибка форматирования: {e}")


def get_full_data_t_api():
    print("🚀 Запуск скрипта для работы с API Т-Банк Инвестиции")
    print("=" * 50)

    # get_bonds(TOKEN, True, bonds_path)
    # get_bonds_data()
    success = get_shares(TOKEN)
    if success:
        get_fundamentals_assets_to_excel(TOKEN)


# Основной запуск
if __name__ == "__main__":
    print("🚀 Запуск скрипта для работы с API Т-Банк Инвестиции")
    print("=" * 50)

    # get_bonds(TOKEN, True, bonds_path)
    get_bonds_data()
    success = get_shares(TOKEN)
    if success:
        get_fundamentals_assets_to_excel(TOKEN)
