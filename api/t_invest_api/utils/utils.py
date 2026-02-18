import time
import pandas as pd
from typing import Optional, List, Dict, Any
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
from t_tech.invest.schemas import Quotation

from constants.config import MAX_RETRIES, RETRY_DELAY


class APIUtils:
    """Утилиты для работы с API"""

    @staticmethod
    def quotation_to_float(q: Quotation) -> float:
        """Преобразует Quotation в float"""
        if not q:
            return 0.0
        return q.units + q.nano / 1e9

    @staticmethod
    def retry_on_rate_limit(func, *args, **kwargs):
        """Повторяет запрос при превышении лимита"""
        for attempt in range(MAX_RETRIES):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                if "лимит" in str(e).lower() and attempt < MAX_RETRIES - 1:
                    print(f"      ⚠️ Лимит API, ждём {RETRY_DELAY} сек...")
                    time.sleep(RETRY_DELAY)
                else:
                    raise
        return None


class ExcelFormatter:
    """Форматирование Excel файлов"""

    @staticmethod
    def format_excel_file(filename: str, df: pd.DataFrame):
        """Форматирует Excel файл"""
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # Стиль заголовков
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # Форматируем заголовки
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_len = 0
                col_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_len = len(str(cell.value))
                        if cell_len > max_len:
                            max_len = cell_len
                    except:
                        pass
                adjusted_width = min(max_len + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Добавляем фильтры
            ws.auto_filter.ref = ws.dimensions

            wb.save(filename)
            print("   ✓ Форматирование применено")

        except Exception as e:
            print(f"   ⚠️ Ошибка форматирования: {e}")

    @staticmethod
    def save_to_excel(data: List[Dict[str, Any]], filename: str):
        """Сохраняет данные в Excel с форматированием"""
        try:
            df = pd.DataFrame(data)

            if df.empty:
                print("❌ Нет данных для сохранения")
                return

            print(f"\n💾 Сохраняю {len(df)} записей в '{filename}'...")
            df.to_excel(filename, index=False, engine="openpyxl")
            ExcelFormatter.format_excel_file(filename, df)

            print(f"✅ Данные успешно сохранены!")
            print(f"📊 Статистика: {len(df)} строк, {len(df.columns)} столбцов")

        except Exception as e:
            print(f"❌ Ошибка сохранения: {e}")
