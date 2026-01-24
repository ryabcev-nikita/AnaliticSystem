import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier, export_text
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, accuracy_score
import warnings
warnings.filterwarnings('ignore')
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# Функции для парсинга данных
def parse_float(value):
    """Парсинг числовых значений"""
    if pd.isna(value) or value == '' or value == '-' or str(value).strip() == '':
        return np.nan
    
    value_str = str(value).strip()
    
    # Если значение уже число
    if isinstance(value, (int, float)):
        return float(value)
    
    # Удаляем пробелы (включая пробелы в числах типа "4 282")
    value_str = value_str.replace(' ', '')
    
    # Заменяем запятые на точки для десятичных дробей
    value_str = value_str.replace(',', '.')
    
    # Удаляем нечисловые символы, кроме минуса и точки
    value_str = re.sub(r'[^\d.-]', '', value_str)
    
    try:
        return float(value_str)
    except:
        # Для случаев типа "-1 885.4"
        try:
            # Пробуем найти число с учетом научной нотации
            match = re.search(r'[-+]?\d*\.?\d+([eE][-+]?\d+)?', value_str)
            if match:
                return float(match.group())
            return np.nan
        except:
            return np.nan

def parse_percent(value):
    """Парсинг процентных значений"""
    if pd.isna(value) or value == '' or value == '-' or str(value).strip() == '':
        return np.nan
    
    value_str = str(value).strip()
    
    # Удаляем пробелы
    value_str = value_str.replace(' ', '')
    
    # Заменяем запятые на точки
    value_str = value_str.replace(',', '.')
    
    # Удаляем знак процента
    value_str = value_str.replace('%', '')
    
    try:
        num = float(value_str)
        # Если число больше 1 (например, 27% как 27), делим на 100
        # Если число меньше 1 (например, 0.27), оставляем как есть
        return num / 100 if num > 1 else num
    except:
        return np.nan

def clean_ticker(value):
    """Очистка тикера от лишних символов"""
    if pd.isna(value) or value == '':
        return ''
    
    ticker = str(value).strip()
    
    # Удаляем лишние символы, оставляем только буквы, цифры и точки
    ticker = re.sub(r'[^a-zA-Z0-9.-]', '', ticker)
    
    return ticker

def load_company_data_from_excel(file_path='companies_data.xlsx'):
    """Загрузка данных о компаниях из Excel файла"""
    try:
        # Загружаем данные из Excel
        df = pd.read_excel(file_path)
        print(f"Файл успешно загружен. Размер: {df.shape}")
        
        # Выводим названия столбцов для отладки
        print(f"Столбцы в файле: {list(df.columns)}")
        
        # Создаем новый DataFrame с правильными названиями столбцов
        companies = []
        
        # Проходим по каждой строке данных
        for idx, row in df.iterrows():
            try:
                # Получаем значения из столбцов
                # Предполагаем, что структура соответствует предоставленным данным
                # Пытаемся определить номера столбцов по их содержимому
                
                # Ищем столбец с названиями компаний
                name_col = None
                ticker_col = None
                cap_col = None
                ev_col = None
                revenue_col = None
                profit_col = None
                
                # Простой способ: если есть русские буквы в названии - это название компании
                for col in df.columns:
                    if isinstance(row[col], str) and any(cyr in row[col] for cyr in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'):
                        name_col = col
                        break
                
                # Ищем тикер (обычно 3-5 латинских букв)
                for col in df.columns:
                    if isinstance(row[col], str) and re.match(r'^[A-Z]{2,5}$', str(row[col]).strip()):
                        ticker_col = col
                        break
                
                # Если не нашли тикер, проверяем столбцы с цифровыми данными
                # Это эвристический подход - в реальности нужно знать структуру файла
                
                # Создаем словарь с данными
                company_data = {
                    '№': idx + 1,
                    'Название': row[name_col] if name_col else f"Компания_{idx+1}",
                    'Тикер': clean_ticker(row[ticker_col]) if ticker_col else '',
                }
                
                # Добавляем числовые данные (попробуем найти их по индексам)
                # Будем использовать индексы столбцов, соответствующие предоставленным данным
                
                # Капитализация (столбец F, индекс 5 в 0-based)
                if len(df.columns) > 5:
                    company_data['Капитализация'] = parse_float(row[df.columns[5]])
                
                # EV (столбец G, индекс 6)
                if len(df.columns) > 6:
                    company_data['EV'] = parse_float(row[df.columns[6]])
                
                # Выручка (столбец H, индекс 7)
                if len(df.columns) > 7:
                    company_data['Выручка'] = parse_float(row[df.columns[7]])
                
                # Чистая прибыль (столбец I, индекс 8)
                if len(df.columns) > 8:
                    company_data['Чистая_прибыль'] = parse_float(row[df.columns[8]])
                
                # ДД ао (столбец J, индекс 9)
                if len(df.columns) > 9:
                    company_data['ДД_ао'] = parse_percent(row[df.columns[9]])
                
                # ДД ап (столбец K, индекс 10)
                if len(df.columns) > 10:
                    company_data['ДД_ап'] = parse_percent(row[df.columns[10]])
                
                # ДД/ЧП (столбец L, индекс 11)
                if len(df.columns) > 11:
                    company_data['ДД_ЧП'] = parse_percent(row[df.columns[11]])
                
                # P/E (столбец M, индекс 12)
                if len(df.columns) > 12:
                    company_data['P_E'] = parse_float(row[df.columns[12]])
                
                # P/S (столбец N, индекс 13)
                if len(df.columns) > 13:
                    company_data['P_S'] = parse_float(row[df.columns[13]])
                
                # P/B (столбец O, индекс 14)
                if len(df.columns) > 14:
                    company_data['P_B'] = parse_float(row[df.columns[14]])
                
                # EV/EBITDA (столбец P, индекс 15)
                if len(df.columns) > 15:
                    company_data['EV_EBITDA'] = parse_float(row[df.columns[15]])
                
                # Рентабельность EBITDA (столбец Q, индекс 16)
                if len(df.columns) > 16:
                    company_data['Рентабельность_EBITDA'] = parse_percent(row[df.columns[16]])
                
                # долг/EBITDA (столбец R, индекс 17)
                if len(df.columns) > 17:
                    company_data['долг_EBITDA'] = parse_float(row[df.columns[17]])
                
                companies.append(company_data)
                
            except Exception as e:
                print(f"Ошибка при обработке строки {idx}: {e}")
                continue
        
        result_df = pd.DataFrame(companies)
        print(f"Обработано {len(result_df)} компаний")
        return result_df
        
    except Exception as e:
        print(f"Ошибка при загрузке файла: {e}")
        return pd.DataFrame()

# Более надежный способ загрузки с известными названиями столбцов
def load_company_data_smart(file_path='companies_data.xlsx'):
    """Умная загрузка данных из Excel файла"""
    try:
        # Загружаем данные из Excel
        df = pd.read_excel(file_path)
        print(f"Файл загружен: {df.shape[0]} строк, {df.shape[1]} столбцов")
        
        # Преобразуем названия столбцов для удобства
        df.columns = [str(col).strip() for col in df.columns]
        
        # Создаем словарь для переименования столбцов
        column_mapping = {}
        
        # Автоматически определяем столбцы по содержимому
        for col in df.columns:
            col_lower = str(col).lower()
            
            # Определяем тип столбца по названию или содержимому
            if any(keyword in col_lower for keyword in ['назван', 'name', 'company']):
                column_mapping[col] = 'Название'
            elif any(keyword in col_lower for keyword in ['тикер', 'ticker', 'symbol']):
                column_mapping[col] = 'Тикер'
            elif any(keyword in col_lower for keyword in ['капит', 'капитал', 'capital', 'рыночн']):
                column_mapping[col] = 'Капитализация'
            elif any(keyword in col_lower for keyword in ['ev', 'enterprise', 'стоимость']):
                column_mapping[col] = 'EV'
            elif any(keyword in col_lower for keyword in ['выруч', 'revenue', 'sales', 'доход']):
                column_mapping[col] = 'Выручка'
            elif any(keyword in col_lower for keyword in ['прибыль', 'profit', 'net', 'чист']):
                column_mapping[col] = 'Чистая_прибыль'
            elif any(keyword in col_lower for keyword in ['дд ао', 'дивиденд', 'dividend']):
                column_mapping[col] = 'ДД_ао'
            elif any(keyword in col_lower for keyword in ['p/e', 'пе', 'price/earnings']):
                column_mapping[col] = 'P_E'
            elif any(keyword in col_lower for keyword in ['p/b', 'пб', 'price/book']):
                column_mapping[col] = 'P_B'
            elif any(keyword in col_lower for keyword in ['рентаб', 'ebitda margin', 'ebitda']):
                column_mapping[col] = 'Рентабельность_EBITDA'
            elif any(keyword in col_lower for keyword in ['ev/ebitda']):
                column_mapping[col] = 'EV_EBITDA'
        
        # Переименовываем столбцы
        df = df.rename(columns=column_mapping)
        
        # Если не нашли нужные столбцы, используем позиционные
        if 'Название' not in df.columns and len(df.columns) > 1:
            df = df.rename(columns={df.columns[1]: 'Название'})
        
        if 'Тикер' not in df.columns and len(df.columns) > 2:
            df = df.rename(columns={df.columns[2]: 'Тикер'})
        
        # Создаем список компаний с правильными данными
        companies = []
        
        for idx, row in df.iterrows():
            try:
                # Пропускаем пустые строки
                if pd.isna(row.get('Название', '')) or str(row.get('Название', '')).strip() == '':
                    continue
                
                company_data = {
                    '№': idx + 1,
                    'Название': str(row.get('Название', f"Компания_{idx+1}")).strip(),
                    'Тикер': clean_ticker(row.get('Тикер', '')),
                }
                
                # Добавляем числовые данные
                numeric_fields = [
                    'Капитализация', 'EV', 'Выручка', 'Чистая_прибыль',
                    'P_E', 'P_S', 'P_B', 'EV_EBITDA'
                ]
                
                for field in numeric_fields:
                    if field in row:
                        company_data[field] = parse_float(row[field])
                    else:
                        company_data[field] = np.nan
                
                # Добавляем процентные данные
                percent_fields = ['ДД_ао', 'ДД_ап', 'ДД_ЧП', 'Рентабельность_EBITDA']
                for field in percent_fields:
                    if field in row:
                        company_data[field] = parse_percent(row[field])
                    else:
                        company_data[field] = np.nan
                
                # Добавляем долг/EBITDA
                if 'долг_EBITDA' in row:
                    company_data['долг_EBITDA'] = parse_float(row['долг_EBITDA'])
                else:
                    company_data['долг_EBITDA'] = np.nan
                
                companies.append(company_data)
                
            except Exception as e:
                print(f"Ошибка при обработке строки {idx}: {e}")
                continue
        
        result_df = pd.DataFrame(companies)
        print(f"Успешно обработано {len(result_df)} компаний")
        return result_df
        
    except Exception as e:
        print(f"Ошибка при загрузке файла: {e}")
        return pd.DataFrame()

# Самый простой способ - загрузка с известной структурой
def load_company_data_simple(file_path='companies_data.xlsx'):
    """Простая загрузка данных с предположением о структуре"""
    try:
        # Загружаем все данные из Excel
        df = pd.read_excel(file_path, header=None)  # Без заголовков
        
        print(f"Загружено {df.shape[0]} строк")
        
        # Создаем список для хранения данных компаний
        companies = []
        
        # Проходим по строкам, начиная с 3-й (после заголовков)
        for i in range(2, len(df)):
            try:
                row = df.iloc[i]
                
                # Пропускаем пустые строки
                if pd.isna(row[1]) or str(row[1]).strip() == '':
                    continue
                
                company_data = {
                    '№': int(row[0]) if pd.notna(row[0]) else i-1,
                    'Название': str(row[1]).strip(),
                    'Тикер': clean_ticker(row[2] if len(row) > 2 else ''),
                    'Капитализация': parse_float(row[5] if len(row) > 5 else np.nan),
                    'EV': parse_float(row[6] if len(row) > 6 else np.nan),
                    'Выручка': parse_float(row[7] if len(row) > 7 else np.nan),
                    'Чистая_прибыль': parse_float(row[8] if len(row) > 8 else np.nan),
                    'ДД_ао': parse_percent(row[9] if len(row) > 9 else np.nan),
                    'ДД_ап': parse_percent(row[10] if len(row) > 10 else np.nan),
                    'ДД_ЧП': parse_percent(row[11] if len(row) > 11 else np.nan),
                    'P_E': parse_float(row[12] if len(row) > 12 else np.nan),
                    'P_S': parse_float(row[13] if len(row) > 13 else np.nan),
                    'P_B': parse_float(row[14] if len(row) > 14 else np.nan),
                    'EV_EBITDA': parse_float(row[15] if len(row) > 15 else np.nan),
                    'Рентабельность_EBITDA': parse_percent(row[16] if len(row) > 16 else np.nan),
                    'долг_EBITDA': parse_float(row[17] if len(row) > 17 else np.nan),
                }
                
                companies.append(company_data)
                
            except Exception as e:
                print(f"Ошибка в строке {i}: {e}")
                continue
        
        result_df = pd.DataFrame(companies)
        print(f"Успешно загружено {len(result_df)} компаний")
        return result_df
        
    except Exception as e:
        print(f"Ошибка при загрузке: {e}")
        return pd.DataFrame()


def create_model_tree_solver():
    # Загружаем данные
    print("="*80)
    print("ЗАГРУЗКА ДАННЫХ О КОМПАНИЯХ")
    print("="*80)

    # Пробуем разные способы загрузки
    df = load_company_data_simple('companies_data.xlsx')

    # Если не получилось, пробуем другой метод
    if len(df) == 0:
        print("\nПробуем альтернативный метод загрузки...")
        df = load_company_data_smart('companies_data.xlsx')

    if len(df) == 0:
        print("Не удалось загрузить данные. Проверьте файл и путь.")
        exit()

    # Обработка данных
    print("\nОбработка данных...")

    # Используем ДД_ап если ДД_ао отсутствует
    df['ДД_используемая'] = df.apply(
        lambda x: x['ДД_ао'] if pd.notna(x['ДД_ао']) else x['ДД_ап'], axis=1
    )

    # Расчет дополнительных показателей
    df['Прибыльность'] = df['Чистая_прибыль'] / df['Выручка'].replace(0, np.nan)
    df['EV_Sales'] = df['EV'] / df['Выручка'].replace(0, np.nan)
    df['Рентабельность_капитала'] = df['Чистая_прибыль'] / df['Капитализация'].replace(0, np.nan)

    # Определяем целевую переменную - категория риска/доходности
    def assign_risk_return_category(row):
        """Назначаем категорию на основе мультипликаторов"""
        
        score = 0
        reasons = []
        
        # Дивидендная доходность (чем выше, тем лучше)
        dd = row['ДД_используемая']
        if pd.notna(dd) and dd > 0:
            if dd > 0.08:  # > 8%
                score += 3
                reasons.append("Высокая ДД")
            elif dd > 0.05:  # 5-8%
                score += 2
                reasons.append("Средняя ДД")
            elif dd > 0.02:  # 2-5%
                score += 1
                reasons.append("Низкая ДД")
        
        # P/E (чем ниже, тем лучше) - только положительный
        pe = row['P_E']
        if pd.notna(pe) and pe > 0:
            if pe < 8:
                score += 3
                reasons.append("Низкий P/E")
            elif pe < 12:
                score += 2
                reasons.append("Умеренный P/E")
            elif pe < 20:
                score += 1
                reasons.append("Приемлемый P/E")
        
        # P/B (чем ниже, тем лучше) - только положительный
        pb = row['P_B']
        if pd.notna(pb) and pb > 0:
            if pb < 1:
                score += 3
                reasons.append("P/B < 1 (недооценка)")
            elif pb < 2:
                score += 2
                reasons.append("Низкий P/B")
            elif pb < 3:
                score += 1
                reasons.append("Умеренный P/B")
        
        # Рентабельность EBITDA (чем выше, тем лучше)
        ebitda_margin = row['Рентабельность_EBITDA']
        if pd.notna(ebitda_margin) and ebitda_margin > 0:
            if ebitda_margin > 0.3:  # > 30%
                score += 3
                reasons.append("Высокая рентабельность")
            elif ebitda_margin > 0.2:  # 20-30%
                score += 2
                reasons.append("Хорошая рентабельность")
            elif ebitda_margin > 0.1:  # 10-20%
                score += 1
                reasons.append("Приемлемая рентабельность")
        
        # EV/EBITDA (чем ниже, тем лучше) - только положительный
        ev_ebitda = row['EV_EBITDA']
        if pd.notna(ev_ebitda) and ev_ebitda > 0:
            if ev_ebitda < 5:
                score += 2
                reasons.append("Низкий EV/EBITDA")
            elif ev_ebitda < 10:
                score += 1
                reasons.append("Умеренный EV/EBITDA")
        
        # Определяем категорию по итоговому счету
        if score >= 10:
            category = 'A: Высокая доходность / Низкий риск'
        elif score >= 7:
            category = 'B: Средняя доходность / Средний риск'
        elif score >= 4:
            category = 'C: Низкая доходность / Высокий риск'
        else:
            category = 'D: Спекулятивная / Очень высокий риск'
        
        return pd.Series([category, score, ', '.join(reasons) if reasons else 'Нет явных преимуществ'])

    df[['Категория', 'Счет_качества', 'Причины']] = df.apply(assign_risk_return_category, axis=1)

    # Подготовка данных для дерева решений
    print("\nПодготовка данных для анализа...")

    features = [
        'ДД_используемая', 'P_E', 'P_S', 'P_B', 'EV_EBITDA', 
        'Рентабельность_EBITDA', 'Прибыльность', 'EV_Sales'
    ]

    # Заполняем пропущенные значения медианой
    df_filled = df[features].copy()
    for col in features:
        if col in df_filled.columns:
            median_val = df_filled[col].median()
            df_filled[col] = df_filled[col].fillna(median_val)

    # Целевая переменная
    target = df['Категория']

    # Разделяем данные на обучающую и тестовую выборки
    if len(df) > 10:
        X_train, X_test, y_train, y_test = train_test_split(
            df_filled, target, test_size=0.2, random_state=42, stratify=target
        )
    else:
        X_train, X_test, y_train, y_test = df_filled, df_filled, target, target

    # Создаем и обучаем дерево решений
    clf = DecisionTreeClassifier(
        max_depth=4,
        min_samples_split=5,
        min_samples_leaf=3,
        random_state=42
    )

    clf.fit(X_train, y_train)

    # Оцениваем качество модели
    y_pred = clf.predict(X_test)

    print("\n" + "="*80)
    print("АНАЛИЗ МОДЕЛИ ДЕРЕВА РЕШЕНИЙ")
    print("="*80)

    if len(df) > 10:
        print(f"Точность модели: {accuracy_score(y_test, y_pred):.2%}")
        print("\nОтчет по классификации:")
        print(classification_report(y_test, y_pred))
    else:
        print("Мало данных для оценки точности модели")

    # Применяем модель ко всем данным
    df['Предсказанная_категория'] = clf.predict(df_filled)

    # Визуализируем правила дерева
    tree_rules = export_text(clf, feature_names=features)
    print("\nПравила дерева решений:")
    print(tree_rules)

    # Анализ распределения компаний по категориям
    print("\n" + "="*80)
    print("РАСПРЕДЕЛЕНИЕ КОМПАНИЙ ПО КАТЕГОРИЯМ")
    print("="*80)
    category_distribution = df['Предсказанная_категория'].value_counts().sort_index()
    for cat, count in category_distribution.items():
        print(f"{cat}: {count} компаний ({count/len(df)*100:.1f}%)")

    # Расчет средних показателей по категориям
    print("\nСредние показатели по категориям:")
    stats_cols = ['ДД_используемая', 'P_E', 'P_B', 'Рентабельность_EBITDA', 'EV_EBITDA', 'Счет_качества']
    category_stats = df.groupby('Предсказанная_категория')[stats_cols].mean().round(3)
    print(category_stats)

    # Функция для получения рекомендаций
    def get_recommendations(df, num_recommendations=20):
        """Получить рекомендации по акциям"""
        
        recommendations_list = []
        
        # 1. Акции категории A (лучшие)
        category_a = df[df['Предсказанная_категория'] == 'A: Высокая доходность / Низкий риск'].copy()
        if len(category_a) > 0:
            category_a['Рейтинг_A'] = (
                category_a['ДД_используемая'].fillna(0) * 0.3 +
                (1 / category_a['P_E'].clip(0.1, 100)) * 0.25 +
                (1 / category_a['P_B'].clip(0.1, 50)) * 0.2 +
                category_a['Рентабельность_EBITDA'].fillna(0) * 0.15 +
                (1 / category_a['EV_EBITDA'].clip(0.1, 50)) * 0.1
            )
            top_a = category_a.nlargest(min(num_recommendations // 2, len(category_a)), 'Рейтинг_A')
            top_a['Тип_рекомендации'] = 'Лучшие (категория A)'
            recommendations_list.append(top_a)
        
        # 2. Акции категории B (хорошие)
        category_b = df[df['Предсказанная_категория'] == 'B: Средняя доходность / Средний риск'].copy()
        if len(category_b) > 0:
            category_b['Рейтинг_B'] = (
                category_b['ДД_используемая'].fillna(0) * 0.25 +
                (1 / category_b['P_E'].clip(0.1, 100)) * 0.25 +
                (1 / category_b['P_B'].clip(0.1, 50)) * 0.2 +
                category_b['Рентабельность_EBITDA'].fillna(0) * 0.2 +
                (1 / category_b['EV_EBITDA'].clip(0.1, 50)) * 0.1
            )
            top_b = category_b.nlargest(min(num_recommendations // 4, len(category_b)), 'Рейтинг_B')
            top_b['Тип_рекомендации'] = 'Хорошие (категория B)'
            recommendations_list.append(top_b)
        
        # 3. Дивидендные акции (вне зависимости от категории)
        dividend_stocks = df[df['ДД_используемая'] > 0.05].copy()
        if len(dividend_stocks) > 0:
            dividend_stocks['Рейтинг_дивиденды'] = (
                dividend_stocks['ДД_используемая'] * 0.6 +
                (1 / dividend_stocks['P_E'].clip(0.1, 100)) * 0.2 +
                dividend_stocks['Рентабельность_EBITDA'].fillna(0) * 0.2
            )
            top_dividend = dividend_stocks.nlargest(min(num_recommendations // 4, len(dividend_stocks)), 'Рейтинг_дивиденды')
            top_dividend['Тип_рекомендации'] = 'Дивидендные'
            recommendations_list.append(top_dividend)
        
        # 4. Акции роста (низкие мультипликаторы)
        growth_stocks = df[
            (df['P_E'] > 0) & 
            (df['P_E'] < 10) & 
            (df['P_B'] > 0) & 
            (df['P_B'] < 1.5) &
            (df['Чистая_прибыль'] > 0)
        ].copy()
        
        if len(growth_stocks) > 0:
            growth_stocks['Рейтинг_рост'] = (
                (1 / growth_stocks['P_E'].clip(0.1, 100)) * 0.4 +
                (1 / growth_stocks['P_B'].clip(0.1, 50)) * 0.3 +
                growth_stocks['Рентабельность_EBITDA'].fillna(0) * 0.3
            )
            top_growth = growth_stocks.nlargest(min(num_recommendations // 4, len(growth_stocks)), 'Рейтинг_рост')
            top_growth['Тип_рекомендации'] = 'Роста (низкие мультипликаторы)'
            recommendations_list.append(top_growth)
        
        # Объединяем все рекомендации
        if recommendations_list:
            recommendations = pd.concat(recommendations_list, ignore_index=True)
            
            # Удаляем дубликаты
            recommendations = recommendations.drop_duplicates(subset=['Тикер', 'Название'])
            
            # Ограничиваем общее количество рекомендаций
            recommendations = recommendations.head(num_recommendations)
            
            # Добавляем общий рейтинг
            recommendations['Общий_рейтинг'] = range(len(recommendations), 0, -1)
            
            # Сортируем по общему рейтингу
            recommendations = recommendations.sort_values('Общий_рейтинг', ascending=False)
            
            return recommendations
        else:
            # Если нет рекомендаций, возвращаем топ по счету качества
            return df.nlargest(num_recommendations, 'Счет_качества')

    # Получаем рекомендации
    print("\n" + "="*80)
    print("ТОП РЕКОМЕНДАЦИИ ПО АКЦИЯМ")
    print("="*80)

    recommendations = get_recommendations(df, 15)

    if len(recommendations) > 0:
        # Форматируем вывод
        display_cols = ['Общий_рейтинг', 'Название', 'Тикер', 'Тип_рекомендации', 
                        'Предсказанная_категория', 'ДД_используемая', 'P_E', 'P_B',
                        'Рентабельность_EBITDA', 'EV_EBITDA', 'Счет_качества', 'Причины']
        
        # Отображаем только существующие колонки
        existing_cols = [col for col in display_cols if col in recommendations.columns]
        display_df = recommendations[existing_cols].copy()
        
        # Форматируем проценты и числа
        def format_percent(x):
            if pd.isna(x):
                return "Н/Д"
            try:
                return f"{float(x):.1%}"
            except:
                return str(x)
        
        def format_float(x):
            if pd.isna(x):
                return "Н/Д"
            try:
                return f"{float(x):.1f}"
            except:
                return str(x)
        
        for col in display_df.columns:
            if col in ['ДД_используемая', 'Рентабельность_EBITDA']:
                display_df[col] = display_df[col].apply(format_percent)
            elif col in ['P_E', 'P_B', 'EV_EBITDA', 'Счет_качества']:
                display_df[col] = display_df[col].apply(format_float)
        
        print(display_df.to_string(index=False))
        
        print(f"\nВсего рекомендаций: {len(recommendations)}")
    else:
        print("Не удалось сформировать рекомендации")

    # Функция для сохранения в Excel с форматированием
    def save_to_excel(full_df, recommendations_df, filename='../../data/акции_рекомендации.xlsx'):
        """Сохранить результаты анализа в Excel файл"""
        
        try:
            # Создаем Excel writer
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. Сохраняем ТОП рекомендации
                if len(recommendations_df) > 0:
                    # Подготавливаем данные для сохранения
                    rec_cols = ['Общий_рейтинг', 'Название', 'Тикер', 'Тип_рекомендации',
                            'Предсказанная_категория', 'ДД_используемая', 'P_E', 'P_B',
                            'EV_EBITDA', 'Рентабельность_EBITDA', 'Выручка',
                            'Чистая_прибыль', 'Капитализация', 'EV', 'Счет_качества', 'Причины']
                    
                    # Выбираем только существующие колонки
                    existing_rec_cols = [col for col in rec_cols if col in recommendations_df.columns]
                    rec_to_save = recommendations_df[existing_rec_cols].copy()
                    
                    # Сохраняем
                    rec_to_save.to_excel(writer, sheet_name='ТОП_рекомендации', index=False)
                
                # 2. Сохраняем все данные с классификацией
                all_cols = ['№', 'Название', 'Тикер', 'Капитализация', 'EV', 'Выручка',
                        'Чистая_прибыль', 'ДД_ао', 'ДД_ап', 'ДД_ЧП', 'P_E', 'P_S',
                        'P_B', 'EV_EBITDA', 'Рентабельность_EBITDA', 'долг_EBITDA',
                        'Предсказанная_категория', 'Счет_качества', 'Причины']
                
                existing_all_cols = [col for col in all_cols if col in full_df.columns]
                full_to_save = full_df[existing_all_cols].copy()
                full_to_save.to_excel(writer, sheet_name='Все_компании', index=False)
                
                # 3. Сохраняем статистику по категориям
                stats = full_df.groupby('Предсказанная_категория').agg({
                    'ДД_используемая': ['count', 'mean', 'median', 'min', 'max'],
                    'P_E': ['mean', 'median', 'min', 'max'],
                    'P_B': ['mean', 'median', 'min', 'max'],
                    'Рентабельность_EBITDA': ['mean', 'median', 'min', 'max'],
                    'EV_EBITDA': ['mean', 'median', 'min', 'max']
                }).round(3)
                
                stats.to_excel(writer, sheet_name='Статистика_категорий')
                
                # Получаем workbook для форматирования
                workbook = writer.book
                
                # Форматируем лист с рекомендациями
                if 'ТОП_рекомендации' in workbook.sheetnames:
                    ws = workbook['ТОП_рекомендации']
                    
                    # Стили для заголовков
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Форматируем заголовки
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    
                    # Автоматическая ширина колонок
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Добавляем условное форматирование для категорий
                    # Категория A - зеленый
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws.conditional_formatting.add(f'E2:E{ws.max_row}',
                        FormulaRule(formula=['$E2="A: Высокая доходность / Низкий риск"'], fill=green_fill)
                    )
                    
                    # Категория B - желтый
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    ws.conditional_formatting.add(f'E2:E{ws.max_row}',
                        FormulaRule(formula=['$E2="B: Средняя доходность / Средний риск"'], fill=yellow_fill)
                    )
                    
                    # Категория C - оранжевый
                    orange_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws.conditional_formatting.add(f'E2:E{ws.max_row}',
                        FormulaRule(formula=['$E2="C: Низкая доходность / Высокий риск"'], fill=orange_fill)
                    )
                    
                    # Категория D - красный
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    ws.conditional_formatting.add(f'E2:E{ws.max_row}',
                        FormulaRule(formula=['$E2="D: Спекулятивная / Очень высокий риск"'], fill=red_fill)
                    )
                    
                    # Добавляем фильтр
                    ws.auto_filter.ref = ws.dimensions
                
                # Добавляем лист с инструкциями
                ws_instructions = workbook.create_sheet("Инструкции")
                instructions = [
                    ["ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ОТЧЕТА"],
                    [""],
                    ["1. ТОП_рекомендации - лучшие акции для инвестирования"],
                    ["   - Сортированы по рейтингу"],
                    ["   - Цветовая маркировка по категориям риска"],
                    ["   - Фильтр для удобной навигации"],
                    [""],
                    ["2. Все_компании - полный список компаний с классификацией"],
                    ["   - Все мультипликаторы и показатели"],
                    ["   - Категория, присвоенная моделью"],
                    [""],
                    ["3. Статистика_категорий - сводные данные по категориям"],
                    ["   - Средние значения мультипликаторов"],
                    ["   - Количество компаний в каждой категории"],
                    [""],
                    ["КАТЕГОРИИ РИСКА/ДОХОДНОСТИ:"],
                    ["A: Высокая доходность / Низкий риск - Лучший выбор для консервативных инвесторов"],
                    ["B: Средняя доходность / Средний риск - Сбалансированный риск/доходность"],
                    ["C: Низкая доходность / Высокий риск - Требуют осторожности, потенциал роста"],
                    ["D: Спекулятивная / Очень высокий риск - Высокорисковые инвестиции"],
                    [""],
                    ["ТИПЫ РЕКОМЕНДАЦИЙ:"],
                    ["Лучшие (категория A) - Компании с лучшими показателями"],
                    ["Хорошие (категория B) - Компании с хорошими показателями"],
                    ["Дивидендные - Высокая дивидендная доходность"],
                    ["Роста (низкие мультипликаторы) - Низкие P/E и P/B, потенциал роста"],
                ]
                
                for i, instruction in enumerate(instructions, 1):
                    ws_instructions.cell(row=i, column=1, value=instruction[0])
                
                # Форматируем заголовок инструкций
                ws_instructions['A1'].font = Font(bold=True, size=14)
                for row in range(1, 4):
                    ws_instructions.row_dimensions[row].height = 25
                ws_instructions.column_dimensions['A'].width = 70
            
            print(f"\nРезультаты сохранены в файл: {filename}")
            print(f"Файл содержит:")
            print(f"1. 'ТОП_рекомендации' - {len(recommendations_df)} лучших акций")
            print(f"2. 'Все_компании' - все {len(full_df)} компаний с классификацией")
            print(f"3. 'Статистика_категорий' - сводные данные")
            print(f"4. 'Инструкции' - руководство по использованию")
            
        except Exception as e:
            print(f"Ошибка при сохранении в Excel: {e}")
            # Сохраняем простым способом
            full_df.to_excel(f'резервная_копия_{filename}', index=False)
            print(f"Создана резервная копия: резервная_копия_{filename}")

    # Запускаем сохранение
    if len(recommendations) > 0:
        save_to_excel(df, recommendations, 'анализ_акций_рекомендации.xlsx')
    else:
        # Если нет рекомендаций, сохраняем хотя бы полные данные
        df.to_excel('анализ_акций.xlsx', index=False)
        print("Сохранены полные данные в 'анализ_акций.xlsx'")

    # Дополнительный анализ
    print("\n" + "="*80)
    print("ДОПОЛНИТЕЛЬНЫЙ АНАЛИЗ")
    print("="*80)

    # Топ-5 по дивидендной доходности
    print("\nТоп-5 по дивидендной доходности:")
    dividend_stocks = df[df['ДД_используемая'] > 0]
    if len(dividend_stocks) > 0:
        top_dd = dividend_stocks.nlargest(5, 'ДД_используемая')
        for _, row in top_dd.iterrows():
            print(f"{row['Название']} ({row['Тикер']}): ДД = {row['ДД_используемая']:.1%}, "
                f"Категория: {row['Предсказанная_категория']}")
    else:
        print("Нет компаний с положительной дивидендной доходностью")

    # Топ-5 по низкому P/E (только положительные)
    print("\nТоп-5 по низкому P/E:")
    positive_pe = df[(df['P_E'] > 0) & (df['P_E'] < 50)]
    if len(positive_pe) > 0:
        top_pe = positive_pe.nsmallest(5, 'P_E')
        for _, row in top_pe.iterrows():
            print(f"{row['Название']} ({row['Тикер']}): P/E = {row['P_E']:.1f}, "
                f"Категория: {row['Предсказанная_категория']}")

    # Топ-5 по рентабельности EBITDA
    print("\nТоп-5 по рентабельности EBITDA:")
    positive_ebitda = df[df['Рентабельность_EBITDA'] > 0]
    if len(positive_ebitda) > 0:
        top_ebitda = positive_ebitda.nlargest(5, 'Рентабельность_EBITDA')
        for _, row in top_ebitda.iterrows():
            print(f"{row['Название']} ({row['Тикер']}): EBITDA margin = {row['Рентабельность_EBITDA']:.1%}, "
                f"Категория: {row['Предсказанная_категория']}")

    # Анализ по отраслям (группировка по первым буквам названий для демонстрации)
    print("\nАнализ распределения по категориям:")
    for category in sorted(df['Предсказанная_категория'].unique()):
        cat_companies = df[df['Предсказанная_категория'] == category]
        print(f"\n{category} ({len(cat_companies)} компаний):")
        for _, company in cat_companies.head(3).iterrows():
            print(f"  - {company['Название']} ({company['Тикер']})")

    print("\n" + "="*80)
    print("АНАЛИЗ ЗАВЕРШЕН")
    print("="*80)
    print(f"Всего проанализировано {len(df)} компаний")
    print(f"Сформировано {len(recommendations)} рекомендаций")
    print(f"Результаты сохранены в Excel файл")


if __name__ == "__main__":
    create_model_tree_solver()